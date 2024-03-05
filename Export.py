from PySide6.QtWidgets import QWidget, QApplication
from qasync import QEventLoop
import sys
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
import ctypes
import os
import psutil
import time

class Export:
    def __init__(self, filename, filtered_array, minimo, maximo, media, varianza, des_tipica):
        self.filename = filename
        self.filtered_array = filtered_array
        self.minimo = minimo
        self.maximo = maximo
        self.media = media
        self.varianza = varianza
        self.des_tipica = des_tipica

    def generate_report(self):
        try:
            filtered_array = pd.DataFrame({
                'Frecuencia Cardiaca': self.filtered_array
            })

            # Crear un DataFrame con los resultados de las estadísticas
            resultados_df = pd.DataFrame({
                'Minimo': [self.minimo],
                'Maximo': [self.maximo],
                'Media': [self.media],
                'Varianza': [self.varianza],
                'Desviacion Tipica': [self.des_tipica]
            })

            # Crear un escritor Excel
            nombre_archivo_excel = 'resultados_estadisticas.xlsx'

            # Intentar cerrar el archivo si está abierto
            for process in psutil.process_iter(['pid', 'name']):
                if process.info['name'] == 'EXCEL.EXE' and nombre_archivo_excel in ' '.join(process.cmdline()):
                    print(f"Cerrando Excel para {nombre_archivo_excel} (PID: {process.info['pid']})...")
                    os.system(f'Taskkill /PID {process.info["pid"]}')
                    time.sleep(2)  # Esperar un poco para asegurarse de que Excel se haya cerrado


            filepath = os.path.join(self.filename, nombre_archivo_excel)

            try:
                with pd.ExcelWriter(filepath, engine='openpyxl', mode='w') as writer:
                    # Registro cardiaco
                    filtered_array.to_excel(writer, sheet_name='Array Original', index=False, header=True)

                    # Statistic Data
                    resultados_df.to_excel(writer, sheet_name='Estadisticas', index=False, header=True, startcol=2)
                    
                    worksheet = writer.sheets['Estadisticas']

                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

                    # Graphic
                    chart = plt.figure(figsize=(6, 4)).add_subplot(111)
                    chart.plot(filtered_array['Frecuencia Cardiaca'])

                    chart.axvline(x=filtered_array['Frecuencia Cardiaca'].idxmax(), color='r', linestyle='--', label=f'Máximo: {self.maximo}')
                    chart.axvline(x=filtered_array['Frecuencia Cardiaca'].idxmin(), color='g', linestyle='--', label=f'Mínimo: {self.minimo}')
                    chart.axhline(y=self.media, color='b', linestyle='--', label=f'Media: {self.media}')

                    chart.legend() 

                    chart.set_xlabel('Tiempo (s)')
                    chart.set_ylabel('Frecuencia Cardiaca (BPM)')

                    image_stream = BytesIO()
                    plt.savefig(image_stream, format='png')
                    plt.close()

                    image = openpyxl.drawing.image.Image(image_stream)

                    top_left_cell = 'C4'  
                    worksheet.add_image(image, top_left_cell)

                    
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

                        

            except PermissionError:
                ctypes.windll.user32.MessageBoxW(0, f"El archivo {nombre_archivo_excel} está abierto en otro programa. Ciérrelo y vuelva a intentarlo.", "Error", 0x10)
                sys.exit(1)

            print(f"Resultados almacenados en {nombre_archivo_excel}")


            print("File saved succesfully!")

        except Exception as e:
            ctypes.windll.user32.MessageBoxW(0, f"Se produjo una excepción: {str(e)}", "Error", 0x10)
            sys.exit(1)

